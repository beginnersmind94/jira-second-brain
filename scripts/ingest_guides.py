#!/usr/bin/env python3
"""Download public guide PDFs into an organized raw/guides/pdf library.

This is Step 1 only: no Markdown extraction, workflow extraction, wiki output,
HTML generation, backend work, or training docs.
"""

import argparse
import csv
import os
import re
import shutil
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

csv.field_size_limit(10**9)

REPO = Path(__file__).resolve().parent.parent
GUIDES_DIR = REPO / "raw" / "guides"
PDF_DIR = GUIDES_DIR / "pdf"
MANIFEST_PATH = GUIDES_DIR / "manifest.csv"
FAILURES_PATH = GUIDES_DIR / "failures.csv"

MANIFEST_FIELDS = [
    "id",
    "title",
    "platform",
    "module",
    "content_type",
    "source_url",
    "local_pdf",
    "status",
    "retrieved_at",
    "sensitive_url",
    "error",
]

FAILURE_FIELDS = [
    "id",
    "title",
    "source_url",
    "error_type",
    "error_message",
    "attempted_at",
    "sensitive_url",
]

OPTIONAL_COLUMNS = {
    "title": "Content Title",
    "platform": "Platform (SC or FH)",
    "module": "Module",
    "content_type": "Content Type",
}

SENSITIVE_QUERY_RE = re.compile(
    r"(x-amz-signature|x-amz-credential|x-amz-security-token|token|credential|"
    r"signature|expires|access[_-]?key|secret|auth|sig)",
    re.IGNORECASE,
)

WINDOWS_ILLEGAL_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]+')


class GuideError(Exception):
    """Expected per-guide failure."""


def repo_rel(path):
    return path.relative_to(REPO).as_posix()


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def clean_cell(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_header(value):
    return re.sub(r"\s+", " ", clean_cell(value)).strip()


def find_column(headers, desired):
    desired_norm = normalize_header(desired).casefold()
    for header in headers:
        if normalize_header(header).casefold() == desired_norm:
            return header
    return None


def slugify(value, fallback):
    value = clean_cell(value) or fallback
    value = value.replace("&", " and ")
    value = WINDOWS_ILLEGAL_RE.sub(" ", value)
    value = re.sub(r"[^A-Za-z0-9._ -]+", " ", value)
    value = re.sub(r"[\s_]+", "-", value.strip())
    value = re.sub(r"-{2,}", "-", value)
    value = value.strip("-. ")
    return value or fallback


def read_csv_input(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return []
        rows = []
        for row in reader:
            rows.append({normalize_header(k): clean_cell(v) for k, v in row.items() if k is not None})
        return rows


def read_excel_input(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl is required for .xlsx/.xlsm input. "
            "Install it with: python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []

    headers = [normalize_header(v) for v in raw_headers]
    rows = []
    for raw_row in rows_iter:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = clean_cell(raw_row[idx] if idx < len(raw_row) else "")
        if any(row.values()):
            rows.append(row)
    return rows


def read_input(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_input(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_input(path)
    if suffix == ".xls":
        raise SystemExit("Unsupported .xls input. Export as .xlsx or .csv, then rerun.")
    raise SystemExit(f"Unsupported input file type: {path.suffix}. Use .csv, .xlsx, or .xlsm.")


def read_csv_dicts(path):
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def atomic_write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=path.name + ".", dir=str(path.parent), text=True)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow({field: row.get(field, "") for field in fieldnames})
        os.replace(tmp_name, path)
    except Exception:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


def sensitive_url(url):
    parsed = urllib.parse.urlsplit(url)
    if not parsed.query:
        return False
    if SENSITIVE_QUERY_RE.search(parsed.query):
        return True
    for key, _value in urllib.parse.parse_qsl(parsed.query, keep_blank_values=True):
        if SENSITIVE_QUERY_RE.search(key):
            return True
    return False


def prepare_input_rows(raw_rows):
    if not raw_rows:
        return []

    headers = list(raw_rows[0].keys())
    url_col = find_column(headers, "Document URL")
    if not url_col:
        raise SystemExit("Missing required column: Document URL")

    optional_cols = {key: find_column(headers, name) for key, name in OPTIONAL_COLUMNS.items()}
    prepared = []
    for raw in raw_rows:
        prepared.append(
            {
                "title": clean_cell(raw.get(optional_cols["title"], "")) if optional_cols["title"] else "",
                "platform": clean_cell(raw.get(optional_cols["platform"], "")) if optional_cols["platform"] else "",
                "module": clean_cell(raw.get(optional_cols["module"], "")) if optional_cols["module"] else "",
                "content_type": clean_cell(raw.get(optional_cols["content_type"], "")) if optional_cols["content_type"] else "",
                "source_url": clean_cell(raw.get(url_col, "")),
            }
        )
    return prepared


def load_manifest():
    by_id = {}
    url_to_id = {}
    for row in read_csv_dicts(MANIFEST_PATH):
        guide_id = row.get("id", "").strip()
        url = row.get("source_url", "").strip()
        if guide_id:
            by_id[guide_id] = {field: row.get(field, "") for field in MANIFEST_FIELDS}
        if guide_id and url:
            url_to_id[url] = guide_id
    return by_id, url_to_id


def load_failures():
    failures = {}
    for row in read_csv_dicts(FAILURES_PATH):
        guide_id = row.get("id", "").strip()
        key = guide_id or row.get("source_url", "").strip() or row.get("title", "").strip()
        if key:
            failures[key] = {field: row.get(field, "") for field in FAILURE_FIELDS}
    return failures


def next_guide_id(existing_rows):
    highest = 0
    for row in existing_rows:
        match = re.fullmatch(r"GUIDE-(\d{3,})", row.get("id", "").strip())
        if match:
            highest = max(highest, int(match.group(1)))
    return f"GUIDE-{highest + 1:03d}"


def local_pdf_path(guide_id, title, platform, module, content_type):
    platform_slug = slugify(platform, "Unknown-Platform")
    module_slug = slugify(module, "Unknown-Module")
    type_slug = slugify(content_type, "Unknown-Content-Type")
    title_slug = slugify(title, "untitled-guide").lower()
    filename = f"{guide_id}-{title_slug}.pdf"
    return PDF_DIR / platform_slug / module_slug / type_slug / filename


def old_flat_pdf_path(guide_id):
    return PDF_DIR / f"{guide_id}.pdf"


def migrate_flat_pdf_if_present(guide_id, target_path, force=False):
    flat_path = old_flat_pdf_path(guide_id)
    if not flat_path.exists() or flat_path.resolve() == target_path.resolve():
        return False
    target_path.parent.mkdir(parents=True, exist_ok=True)
    if target_path.exists() and not force:
        return False
    if target_path.exists() and force:
        target_path.unlink()
    shutil.move(str(flat_path), str(target_path))
    return True


def download_pdf(url, dest, timeout):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "jira-brain-guide-download/1.0",
            "Accept": "application/pdf,*/*;q=0.8",
        },
    )
    dest.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=dest.name + ".", suffix=".tmp", dir=str(dest.parent))
    try:
        bytes_written = 0
        with os.fdopen(fd, "wb") as tmp:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                status = getattr(response, "status", 200)
                if status >= 400:
                    raise GuideError(f"HTTP {status}")
                while True:
                    chunk = response.read(1024 * 256)
                    if not chunk:
                        break
                    tmp.write(chunk)
                    bytes_written += len(chunk)
        if bytes_written == 0:
            raise GuideError("Download returned an empty response")
        os.replace(tmp_name, dest)
    except (urllib.error.URLError, TimeoutError) as exc:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise GuideError(f"Download failed: {exc}") from exc
    except Exception:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise


def log_name(row):
    return f"{row['id']} - {row['title'] or '(untitled)'}"


def manifest_sort_key(row):
    match = re.search(r"\d+", row.get("id", ""))
    return int(match.group(0)) if match else 0


def main():
    parser = argparse.ArgumentParser(description="Download public SchoolCafe guide PDFs into raw/guides/pdf/.")
    parser.add_argument("input_file", help="Path to .csv, .xlsx, or .xlsm guide masterlist.")
    parser.add_argument("--force", action="store_true", help="Redownload PDFs even when the organized file exists.")
    parser.add_argument("--timeout", type=int, default=60, help="Download timeout in seconds per guide.")
    args = parser.parse_args()

    input_path = Path(args.input_file).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    PDF_DIR.mkdir(parents=True, exist_ok=True)

    input_rows = prepare_input_rows(read_input(input_path))
    manifest_by_id, url_to_id = load_manifest()
    failures = load_failures()
    existing_rows_for_ids = list(manifest_by_id.values())

    total = len(input_rows)
    downloaded = 0
    skipped_existing = 0
    failed = 0

    for idx, input_row in enumerate(input_rows, 1):
        attempted_at = now_iso()
        url = input_row["source_url"]
        title = input_row["title"]

        if not url:
            failed += 1
            key = f"ROW-{idx:03d}"
            failures[key] = {
                "id": "",
                "title": title,
                "source_url": "",
                "error_type": "missing_url",
                "error_message": "Document URL is blank",
                "attempted_at": attempted_at,
                "sensitive_url": "false",
            }
            print(f"ROW-{idx:03d} - {title or '(untitled)'}: failed: missing Document URL")
            continue

        guide_id = url_to_id.get(url)
        if not guide_id:
            guide_id = next_guide_id(existing_rows_for_ids)
            existing_rows_for_ids.append({"id": guide_id})
            url_to_id[url] = guide_id

        existing = manifest_by_id.get(guide_id, {})
        row = {
            "id": guide_id,
            "title": title or existing.get("title", ""),
            "platform": input_row["platform"] or existing.get("platform", ""),
            "module": input_row["module"] or existing.get("module", ""),
            "content_type": input_row["content_type"] or existing.get("content_type", ""),
            "source_url": url,
            "local_pdf": "",
            "status": "",
            "retrieved_at": attempted_at,
            "sensitive_url": str(sensitive_url(url)).lower(),
            "error": "",
        }
        target_pdf = local_pdf_path(row["id"], row["title"], row["platform"], row["module"], row["content_type"])
        row["local_pdf"] = repo_rel(target_pdf)

        try:
            migrated = migrate_flat_pdf_if_present(row["id"], target_pdf, force=args.force)
            if migrated:
                print(f"{log_name(row)}: moved old flat PDF into organized path")

            if target_pdf.exists() and target_pdf.stat().st_size > 0 and not args.force:
                row["status"] = "skipped_existing"
                row["retrieved_at"] = existing.get("retrieved_at") or attempted_at
                skipped_existing += 1
                print(f"{log_name(row)}: skipped existing PDF")
            else:
                print(f"{log_name(row)}: downloading PDF")
                download_pdf(url, target_pdf, args.timeout)
                row["status"] = "downloaded"
                downloaded += 1

            failures.pop(row["id"], None)
            failures.pop(row["source_url"], None)
            row["error"] = ""
        except GuideError as exc:
            failed += 1
            row["status"] = "failed"
            row["error"] = str(exc)
            failures[row["id"]] = {
                "id": row["id"],
                "title": row["title"],
                "source_url": row["source_url"],
                "error_type": exc.__class__.__name__,
                "error_message": str(exc),
                "attempted_at": attempted_at,
                "sensitive_url": row["sensitive_url"],
            }
            print(f"{log_name(row)}: failed: {exc}")

        manifest_by_id[row["id"]] = row

    manifest_rows = sorted(manifest_by_id.values(), key=manifest_sort_key)
    failure_rows = sorted(failures.values(), key=lambda row: row.get("id", ""))

    atomic_write_csv(MANIFEST_PATH, manifest_rows, MANIFEST_FIELDS)
    atomic_write_csv(FAILURES_PATH, failure_rows, FAILURE_FIELDS)

    print("")
    print("Guide PDF download summary")
    print(f"  total rows: {total}")
    print(f"  downloaded: {downloaded}")
    print(f"  skipped existing: {skipped_existing}")
    print(f"  failed: {failed}")
    print(f"  pdf root: {repo_rel(PDF_DIR)}")

    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
