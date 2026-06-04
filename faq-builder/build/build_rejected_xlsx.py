"""Build the rejected-tickets Excel register (deliverables/) from the workflow result JSON.
Summary sheet is a funnel: total unique tickets -> combined into N questions -> rejected."""
import json, os, re, collections, runpy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BUILD = os.path.dirname(os.path.abspath(__file__))
PROJ = os.path.dirname(BUILD)
SPECS = os.path.join(BUILD, "specs")
CONV = os.path.join(PROJ, "conversations")
REJECTED_JSON = os.path.join(BUILD, "audit", "_rejected_final.json")
OUT = os.path.join(PROJ, "deliverables", "PrimeroEdge-FAQ-Rejected-Tickets.xlsx")
URL = "https://primeroedge.freshdesk.com/a/tickets/{}"

# Document -> (source builder, filter jsonl). builder is ('spec', name) or ('py', filename).
DOCS = [
    ("Menu Cycles",         ("py", "make_menu_cycles_faq.py"),      "conversations.jsonl"),
    ("Ingredients",         ("py", "make_ingredients_faq.py"),      "conversations_ingredients.jsonl"),
    ("Transfer Items",      ("spec", "transfer_items"),             "conversations_transfer_items.jsonl"),
    ("Assign Menus",        ("spec", "assign_menus"),               "conversations_assign_menus.jsonl"),
    ("Perpetual Inventory", ("spec", "perpetual_inventory"),        "conversations_perpetual_inventory.jsonl"),
    ("Receipt w/o Order",   ("spec", "receipt_wo_order"),           "conversations_receipt_wo_order.jsonl"),
    ("Production Orders",   ("spec", "production_orders"),          "conversations_production_orders.jsonl"),
]


def sections_of(builder):
    kind, ref = builder
    if kind == "spec":
        ns = runpy.run_path(os.path.join(SPECS, f"{ref}_spec.py"))
    else:
        ns = {"__name__": "x", "__file__": os.path.join(PROJ, ref)}
        exec(open(os.path.join(PROJ, ref), encoding="utf-8").read(), ns)
    return ns["SECTIONS"]


def filter_ids(jl):
    return set(str(json.loads(l)["ticket_id"]) for l in open(os.path.join(CONV, jl), encoding="utf-8") if l.strip())


rows = json.load(open(REJECTED_JSON, encoding="utf-8"))
rows.sort(key=lambda r: (r["document"], r["ticket_id"]))
rej_by_doc = collections.Counter(r["document"] for r in rows)

NAVY = "1F3A5F"; LIGHT = "EAF1F3"
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", fgColor=NAVY)
wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")
ctr = Alignment(vertical="center", horizontal="center")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ---- Sheet 1: Rejected Tickets (detail) ----
ws = wb.active
ws.title = "Rejected Tickets"
cols = [("Document", 20), ("Ticket ID", 11), ("Freshdesk URL", 42), ("Subject", 46),
        ("Date", 12), ("Rejection Category", 22), ("Why rejected", 60), ("Citation (verbatim)", 64)]
for i, (name, width) in enumerate(cols, 1):
    c = ws.cell(1, i, name); c.font = hf; c.fill = hfill
    c.alignment = Alignment(vertical="center", wrap_text=True); c.border = border
    ws.column_dimensions[get_column_letter(i)].width = width
for r, row in enumerate(rows, 2):
    url = URL.format(row["ticket_id"])
    vals = [row["document"], row["ticket_id"], url, row["subject"], row["date"],
            row["category"], row["reason"], row["citation"]]
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v); c.border = border
        c.alignment = top if i in (1, 2, 5, 6) else wrap
    ws.cell(r, 3).hyperlink = url; ws.cell(r, 3).font = Font(color="2E7D8A", underline="single")
    if r % 2 == 0:
        for i in range(1, len(cols) + 1):
            if ws.cell(r, i).fill.patternType is None:
                ws.cell(r, i).fill = PatternFill("solid", fgColor=LIGHT)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

# ---- Sheet 2: Coverage funnel ----
s = wb.create_sheet("Coverage Summary")
s["A1"] = "Coverage funnel: total unique tickets  ->  combined into questions  +  rejected"
s["A1"].font = Font(bold=True, size=12, color=NAVY)
hdr = ["Document", "Total unique tickets", "Tickets used", "Combined into (questions)", "Rejected"]
widths = [22, 20, 14, 24, 12]
for i, (h, w) in enumerate(zip(hdr, widths), 1):
    c = s.cell(3, i, h); c.font = hf; c.fill = hfill
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center"); c.border = border
    s.column_dimensions[get_column_letter(i)].width = w

tot_unique = tot_used = tot_q = tot_rej = 0
r = 4
for name, builder, jl in DOCS:
    secs = sections_of(builder)
    questions = sum(len(qas) for _, qas in secs)
    cited = set()
    for _, qas in secs:
        for q, a, src in qas:
            cited.update(re.findall(r"\d{5,7}", src))
    total = len(filter_ids(jl))
    rej = rej_by_doc.get(name, 0)
    used = total - rej
    vals = [name, total, used, questions, rej]
    for i, v in enumerate(vals, 1):
        c = s.cell(r, i, v); c.border = border
        c.alignment = top if i == 1 else ctr
    tot_unique += total; tot_used += used; tot_q += questions; tot_rej += rej
    r += 1
for i, v in enumerate(["TOTAL", tot_unique, tot_used, tot_q, tot_rej], 1):
    c = s.cell(r, i, v); c.font = Font(bold=True); c.border = border
    c.alignment = top if i == 1 else ctr
    c.fill = PatternFill("solid", fgColor=LIGHT)

note_r = r + 2
s.cell(note_r, 1, "How to read this: every ticket in a function's Freshdesk filter (Total unique) is either "
                  "USED in the FAQ - consolidated into the listed number of questions (one question often "
                  "combines several tickets) - or REJECTED (no distinct, reusable guidance) and itemized on "
                  "the 'Rejected Tickets' sheet. Total unique = Tickets used + Rejected.").alignment = wrap
s.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=5)
s.row_dimensions[note_r].height = 60

# by-category mini table
cr = note_r + 2
s.cell(cr, 1, "Rejected tickets by category").font = Font(bold=True, size=12, color=NAVY)
s.cell(cr + 1, 1, "Category").font = hf; s.cell(cr + 1, 1).fill = hfill
s.cell(cr + 1, 2, "Count").font = hf; s.cell(cr + 1, 2).fill = hfill
rr = cr + 2
for cat, n in sorted(collections.Counter(x["category"] for x in rows).items(), key=lambda x: -x[1]):
    s.cell(rr, 1, cat); s.cell(rr, 2, n); rr += 1

try:
    wb.save(OUT)
    print("Wrote", OUT)
except PermissionError:
    alt = OUT.replace(".xlsx", "-RECONCILED.xlsx")
    wb.save(alt)
    print("LOCKED (file open). Wrote", alt, "-- close the original and rerun to overwrite the canonical name.")
print("funnel: total_unique=%d used=%d questions=%d rejected=%d" % (tot_unique, tot_used, tot_q, tot_rej))
