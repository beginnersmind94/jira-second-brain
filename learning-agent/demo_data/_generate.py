"""Generate the Demo ISD CSV + XLSX roster for the Projects flow demo.

Run from the learning-agent dir:
    python demo_data/_generate.py

Emits:
    demo_data/demo_isd_sites.csv   (8 rows)
    demo_data/demo_isd_users.csv   (60 rows)
    demo_data/demo_isd_roster.xlsx (two sheets: Sites, Users)

Distribution: 3 CN Director, 9 Site Manager, 48 Cashier across 8 sites.
Roles match the app's existing canonical enum (Cashier / Site Manager / CN Director).
"""
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent
SITES_CSV = HERE / "demo_isd_sites.csv"
USERS_CSV = HERE / "demo_isd_users.csv"
XLSX = HERE / "demo_isd_roster.xlsx"

SITES = [
    ("Lincoln Elementary", "1200 Lincoln Ave, Demo City, TX 77001"),
    ("Roosevelt Elementary", "850 Roosevelt St, Demo City, TX 77002"),
    ("Jefferson Middle School", "2400 Jefferson Blvd, Demo City, TX 77003"),
    ("Madison Middle School", "1700 Madison Way, Demo City, TX 77004"),
    ("Hamilton High School", "3300 Hamilton Pkwy, Demo City, TX 77005"),
    ("Franklin High School", "5500 Franklin Rd, Demo City, TX 77006"),
    ("Adams K-8 Academy", "900 Adams Dr, Demo City, TX 77007"),
    ("Carver Early Learning Center", "410 Carver Ln, Demo City, TX 77008"),
]

# (role, count, sites pool to round-robin across)
# CN Directors are district-wide; we anchor them at Hamilton (the district seat) so
# every row has a valid site_name (the importer requires it).
CN_DIRECTORS = 3
SITE_MGR_PER_SITE = {  # 9 total: 1 per site, +1 extra at Hamilton HS
    "Lincoln Elementary": 1, "Roosevelt Elementary": 1,
    "Jefferson Middle School": 1, "Madison Middle School": 1,
    "Hamilton High School": 2, "Franklin High School": 1,
    "Adams K-8 Academy": 1, "Carver Early Learning Center": 1,
}
CASHIERS_PER_SITE = {  # 48 total, distributed by school size
    "Lincoln Elementary": 4, "Roosevelt Elementary": 4,
    "Jefferson Middle School": 5, "Madison Middle School": 5,
    "Hamilton High School": 9, "Franklin High School": 8,
    "Adams K-8 Academy": 7, "Carver Early Learning Center": 6,
}

# Deterministic name pool — keeps the file stable across regenerations.
FIRST_NAMES = [
    "Maria", "James", "Aisha", "Carlos", "Linda", "Wei", "Diego", "Sarah",
    "Robert", "Priya", "Kenji", "Grace", "Marcus", "Elena", "Tyler", "Fatima",
    "Rosa", "Andre", "Nia", "Hector", "Joy", "Samuel", "Lucia", "Omar",
    "Beth", "Trang", "Carl", "Imani", "Naomi", "Liam", "Sofia", "Eli",
    "Maya", "Jorge", "Hannah", "Vikram", "Adaeze", "Cole", "Mei", "Yara",
    "Brandon", "Theo", "Renee", "Kofi", "Sabrina", "Daniel", "Mira", "Connor",
    "Olive", "Aaron", "Riya", "Chen", "Esther", "Hugo", "Ines", "Felix",
    "Lila", "Owen", "Ada", "Jonah",
]
LAST_NAMES = [
    "Garcia", "Johnson", "Patel", "Nguyen", "Smith", "Williams", "Brown",
    "Lee", "Martinez", "Davis", "Khan", "Lopez", "Wilson", "Thomas", "Reyes",
    "Okafor", "Cohen", "Tran", "Flores", "Bell", "Ramirez", "Young", "Diaz",
    "Foster", "Hernandez", "Anderson", "Walker", "Mitchell", "Carter", "Robinson",
    "Perez", "Sanchez", "Singh", "Ali", "Park", "Kim", "Adams", "Wright",
    "Hill", "Bailey", "Cooper", "Howard", "Bennett", "Murphy", "Watson", "Brooks",
    "Sullivan", "Morales", "Russell", "Gomez", "Castillo", "Romero", "Hughes",
    "Price", "Long", "Foster", "Ortiz", "Stone", "Vargas", "Hayes",
]


def gen_users():
    rows = []
    eid = 1001
    fi = 0
    li = 0

    def take_name():
        nonlocal fi, li
        f = FIRST_NAMES[fi % len(FIRST_NAMES)]
        l = LAST_NAMES[li % len(LAST_NAMES)]
        fi += 1
        li += 3  # stride to keep first/last pairings from repeating in obvious patterns
        return f, l

    def row(role, site_name):
        nonlocal eid
        f, l = take_name()
        e_id = f"E{eid}"
        eid += 1
        email = f"{f.lower()}.{l.lower()}@demoisd.org"
        rows.append({
            "employee_id": e_id,
            "full_name": f"{f} {l}",
            "email": email,
            "role": role,
            "site_name": site_name,
        })

    # CN Directors — district-level, anchored at Hamilton HS
    for _ in range(CN_DIRECTORS):
        row("CN Director", "Hamilton High School")
    # Site Managers
    for site, n in SITE_MGR_PER_SITE.items():
        for _ in range(n):
            row("Site Manager", site)
    # Cashiers
    for site, n in CASHIERS_PER_SITE.items():
        for _ in range(n):
            row("Cashier", site)

    return rows


def write_csv(path, fieldnames, rows):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(path, sites, users):
    wb = openpyxl.Workbook()
    ws_s = wb.active
    ws_s.title = "Sites"
    ws_s.append(["site_name", "address"])
    for s in sites:
        ws_s.append([s[0], s[1]])

    ws_u = wb.create_sheet("Users")
    ws_u.append(["employee_id", "full_name", "email", "role", "site_name"])
    for u in users:
        ws_u.append([u["employee_id"], u["full_name"], u["email"], u["role"], u["site_name"]])

    # Light styling — bold header, header fill, frozen first row, autosized columns.
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3A2A20")  # warm dark like the app
    for ws in (ws_s, ws_u):
        for c in ws[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 42)

    wb.save(path)


def main():
    users = gen_users()
    assert len(SITES) == 8, "expected 8 sites"
    assert len(users) == 60, f"expected 60 users, got {len(users)}"

    write_csv(SITES_CSV, ["site_name", "address"],
              [{"site_name": s[0], "address": s[1]} for s in SITES])
    write_csv(USERS_CSV, ["employee_id", "full_name", "email", "role", "site_name"], users)
    write_xlsx(XLSX, SITES, users)

    by_role = {}
    for u in users:
        by_role[u["role"]] = by_role.get(u["role"], 0) + 1
    print(f"wrote {SITES_CSV.name}: {len(SITES)} rows")
    print(f"wrote {USERS_CSV.name}: {len(users)} rows  by_role={by_role}")
    print(f"wrote {XLSX.name}: Sites + Users sheets")


if __name__ == "__main__":
    main()
